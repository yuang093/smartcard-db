import io
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.core.database import get_db
from app.models.user import User
from app.models.card import Card
from app.api.auth.router import get_current_user

router = APIRouter(tags=["cards-export"])


@router.get("/export")
async def export_cards_xlsx(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """匯出所有名片為 Excel 檔案"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Get all cards for current user
        result = await db.execute(
            select(Card).where(Card.user_id == current_user.id).order_by(Card.created_at.desc())
        )
        cards = result.scalars().all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "名片資料"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        headers = ["Name", "Company", "Title", "Phone", "Mobile", "Email", "Address", "Created"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data
        for row, card in enumerate(cards, 2):
            ws.cell(row=row, column=1, value=str(card.name) if card.name else "").border = thin_border
            ws.cell(row=row, column=2, value=str(card.company) if card.company else "").border = thin_border
            ws.cell(row=row, column=3, value=str(card.title) if card.title else "").border = thin_border
            ws.cell(row=row, column=4, value=str(card.phone) if card.phone else "").border = thin_border
            ws.cell(row=row, column=5, value=str(card.mobile) if card.mobile else "").border = thin_border
            ws.cell(row=row, column=6, value=str(card.email) if card.email else "").border = thin_border
            ws.cell(row=row, column=7, value=str(card.address) if card.address else "").border = thin_border
            ws.cell(row=row, column=8, value=card.created_at.strftime("%Y-%m-%d %H:%M") if card.created_at else "").border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15  # Name
        ws.column_dimensions['B'].width = 20  # Company
        ws.column_dimensions['C'].width = 15  # Title
        ws.column_dimensions['D'].width = 15  # Phone
        ws.column_dimensions['E'].width = 15  # Mobile
        ws.column_dimensions['F'].width = 25  # Email
        ws.column_dimensions['G'].width = 30  # Address
        ws.column_dimensions['H'].width = 18  # Created
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"cards_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"匯出失敗: {str(e)}")
